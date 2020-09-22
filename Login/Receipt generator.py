from tkinter import*
from PIL import ImageTk,Image
import time
import openpyxl
import smtplib
from email.mime import *
import os
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import random
import tkinter.messagebox
b=0;k=0

def checker():
    global b,k
    global reclist
    reclist=[]
    def funmail():

        #Reciept number generator
        recinfo=random.randint(100000,999999)
        while True:
            if recinfo in reclist:
                recinfo=random.randint(100000,999999)
            else:
                break
        mailist=[emailinfo,'electicider123@gmail.com']
        #emailer
        for email_address in mailist:

            name_of_recipeint=nameinfo
            my_address = 'zehratuzfatima2000@gmail.com'
            password = 'Maymoona123'
            my_server = smtplib.SMTP('smtp.gmail.com', 587)
            my_server.starttls()
            my_server.login(my_address, password)
            reciept = MIMEMultipart()
            reciept['From'] = my_address
            reciept['To'] =name_of_recipeint
            reciept['Subject'] = 'Your voting receipt'
            text = 'Email:{} CMS ID:{} Reciept No:{} '.format(emailinfo,cmsinfo,recinfo)
            reclist.append(recinfo)
            reciept.attach(MIMEText(text, 'plain'))
            my_server.sendmail(my_address, email_address, str(reciept))
            tkinter.messagebox.showinfo('Acknowledgment','Receipt sent through mail')
            my_server.quit()
    global nameinfo,cmsinfo,voterinfo,nicinfo,sectinfo,emailinfo,recinfo
    try:
        nameinfo=Name.get()
        cmsinfo=cms.get()
        voterinfo=voter.get()
        nicinfo=nic.get()
        sectinfo=sect.get()
        emailinfo=mail.get()
    except:
        nameinfo='';cmsinfo='';voterinfo='';nicinfo='';sectinfo='';emailinfo=''

    wb=openpyxl.load_workbook('Project1.xlsx')
    sheet=wb['Sheet1']
    rowh=sheet.max_row
    columnh=sheet.max_column
    print(rowh,columnh)
    global a;global i,b,k
    a=0
    list1=[cmsinfo,nicinfo,voterinfo,emailinfo]
    length=len(list1)
    global list2
    if b==0 and k==0:
        list2=[]
    counter=0
    global truth
    for item in range(2,rowh+1):
        counter=0
        a=0
        for i in range(2,columnh+1) :
            c=sheet.cell(column=i,row=item)
            if c.value==list1[a] and (c.value not in list2):
                truth=True

                counter=counter+1
                if counter==4:
                    break
            else:
                truth=False
            a=a+1
        if counter==4:
            list2.append(cmsinfo);list2.append(nicinfo);list2.append(voterinfo);list2.append(emailinfo)
            b+=1;k+=1
            funmail()


    else:
        tkinter.messagebox.showinfo('NOT IDENTIFIED','Please try again\n Unknown or duplicate credentials')









def click1(event):
    global Name
    for i in range(1):
        entry1.config(textvariable=Name)

def hoveron1(event):
    lab1.config(image=namec)
    lab1.place(relx=0.001)
    entry1.config(bg="#E6E7E8")
    global xspeed1
    global yspeed1
    xspeed1 = 0.5
    yspeed1 = 0

    while True:
        try:
            can1.move(rect1, xspeed1, yspeed1)
        except Exception:
            break
        pos1 = can1.coords(rect1)
        if pos1[2] > 10:
            xspeed1 = 0
        root.update()
        time.sleep(0.01)

def hoveroff1(event):
    lab1.config(image=namei)
    lab1.place(relx=0)
    entry1.config(bg="white")
    global xspeed1
    global yspeed1
    xspeed1 = -0.5
    yspeed1 = 0

    while True:
        try:
            can1.move(rect1, xspeed1, yspeed1)
        except Exception:
            break
        pos1 = can1.coords(rect1)
        if pos1[2] < -5:
            xspeed1 = 0
        root.update()
        time.sleep(0.01)


def click2(event):
    global cms
    for i in range(1):
        entry2.config(textvariable=cms)

def hoveron2(event):
    lab2.config(image=cmsc)
    lab2.place(rely=0.05)
    entry2.config(bg="#E6E7E8")
    global xspeed2
    global yspeed2
    xspeed2 = 0.5
    yspeed2 = 0

    while True:
        try:
            can2.move(rect2, xspeed2, yspeed2)
        except Exception:
            break
        pos2 = can2.coords(rect2)
        if pos2[2] > 10:
            xspeed2 = 0
        root.update()
        time.sleep(0.01)


def hoveroff2(event):
    lab2.config(image=cmsi)
    lab2.place(rely=0)
    entry2.config(bg="white")
    global xspeed2
    global yspeed2
    xspeed2 = -0.5
    yspeed2 = 0

    while True:
        try:
            can2.move(rect2, xspeed2, yspeed2)
        except Exception:
            break
        pos2 = can2.coords(rect2)
        if pos2[2] < -5:
            xspeed2 = 0
        root.update()
        time.sleep(0.01)

def click3(event):
    global nic
    for i in range(1):
        entry3.config(textvariable=nic)

def hoveron3(event):
    lab3.config(image=cnicc)
    entry3.config(bg="#E6E7E8")
    global xspeed3
    global yspeed3
    xspeed3 = 0.5
    yspeed3 = 0

    while True:
        try:
            can3.move(rect3, xspeed3, yspeed3)
        except Exception:
            break
        pos3 = can3.coords(rect3)
        if pos3[2] > 10:
            xspeed3 = 0
        root.update()
        time.sleep(0.01)


def hoveroff3(event):
    lab3.config(image=cnici)
    entry3.config(bg="white")
    global xspeed3
    global yspeed3
    xspeed3 = -0.5
    yspeed3 = 0

    while True:
        try:
            can3.move(rect3, xspeed3, yspeed3)
        except Exception:
            break
        pos3 = can3.coords(rect3)
        if pos3[2] < -5:
            xspeed3 = 0
        root.update()
        time.sleep(0.01)


def click5(event):
    global voter
    for i in range(1):
        entry5.config(textvariable=voter)

def hoveron5(event):
    lab5.config(image=voterc)
    entry5.config(bg="#E6E7E8")
    global xspeed5
    global yspeed5
    xspeed5 = 0.5
    yspeed5 = 0

    while True:
        try:
            can5.move(rect5, xspeed5, yspeed5)
        except Exception:
            break
        pos5 = can5.coords(rect5)
        if pos5[2] > 10:
            xspeed5 = 0
        root.update()
        time.sleep(0.01)

def hoveroff5(event):
    lab5.config(image=voteri)
    entry5.config(bg="white")
    global xspeed5
    global yspeed5
    xspeed5 = -0.5
    yspeed5 = 0

    while True:
        try:
            can5.move(rect5, xspeed5, yspeed5)
        except Exception:
            break
        pos5 = can5.coords(rect5)
        if pos5[2] < -5:
            xspeed5 = 0
        root.update()
        time.sleep(0.01)

def click4(event):
    global sect
    for i in range(1):
        entry4.config(textvariable=sect)

def hoveron4(event):
    lab4.config(image=sectionc)
    entry4.config(bg="#E6E7E8")
    global xspeed4
    global yspeed4
    xspeed4 = 0.5
    yspeed4 = 0

    while True:
        try:
            can4.move(rect4, xspeed4, yspeed4)
        except Exception:
            break
        pos4 = can4.coords(rect4)
        if pos4[2] > 10:
            xspeed4 = 0
        root.update()
        time.sleep(0.01)


def hoveroff4(event):
    lab4.config(image=sectioni)
    entry4.config(bg="white")
    global xspeed4
    global yspeed4
    xspeed4 = -0.5
    yspeed4 = 0

    while True:
        try:
            can4.move(rect4, xspeed4, yspeed4)
        except Exception:
            break
        pos4 = can4.coords(rect4)
        if pos4[2] < -5:
            xspeed4 = 0
        root.update()
        time.sleep(0.01)


def click6(event):
    global mail
    for i in range(1):
        entry6.config(textvariable=mail)

def hoveron6(event):
    lab6.config(image=emailc)
    lab6.place(relx=0.001)
    entry6.config(bg="#E6E7E8")
    global xspeed6
    global yspeed6
    xspeed6 = 0.5
    yspeed6 = 0

    while True:
        try:
            can6.move(rect6, xspeed6, yspeed6)
        except Exception:
            break
        pos6 = can6.coords(rect6)
        if pos6[2] > 10:
            xspeed6 = 0
        root.update()
        time.sleep(0.01)


def hoveroff6(event):
    lab6.config(image=emaili)
    lab6.place(relx=0)
    entry6.config(bg="white")
    global xspeed6
    global yspeed6
    xspeed6 = -0.5
    yspeed6 = 0

    while True:
        try:
            can6.move(rect6, xspeed6, yspeed6)
        except Exception:
            break
        pos6 = can6.coords(rect6)
        if pos6[2] < -5:
            xspeed6 = 0
        root.update()
        time.sleep(0.01)

def line():
    global xspeeda
    global yspeeda
    xspeeda = -5
    yspeeda = 0
    while True:
        try:
            canb.move(recta, xspeeda, yspeeda)
        except Exception:
            break
        posa = canb.coords(recta)
        if posa[2] < 1500:
            xspeeda = 0
        root.update()
        time.sleep(0.00001)


def clicke(event):
    button1.config(image=enterc)
    return
def releasee(event):
    button1.config(image=enterh)
    return
def hoverone4(event):
    button1.config(image=enterh)
    return
def hoveroffe4(event):
    button1.config(image=enteri)
    return


def show(var1,var2,var3,var4,var5,var6):
    print(var1)
    print(var2)
    print(var3)
    print(var4)
    print(var5)
    print(var6)
    global Name
    global cms
    global nic
    global sect
    global voter
    global mail

    Name = StringVar()
    cms = IntVar()
    nic = IntVar()
    sect = StringVar()
    voter = IntVar()
    mail = StringVar()

    entry1.config(textvariable=Name)
    entry2.config(textvariable=cms)
    entry3.config(textvariable=nic)
    entry4.config(textvariable=sect)
    entry5.config(textvariable=voter)
    entry6.config(textvariable=mail)


def mains():
    global hovname
    global hovcms
    global hovvoter
    global hovsect
    global hovcnic
    global hovemail
    global root
    global button1
    global Name;global cms;global nic;global sect;global voter;global mail; global wname;global wcms;global wnic;global wsect;global wvoter;global wmail
    global lab1; global lab2; global lab3; global lab4; global lab5; global lab6
    global entry1; global entry2; global entry3; global entry4; global entry5; global entry6
    global can1; global can2; global can3; global can4; global can5; global can6
    global rect1; global rect2; global rect3; global rect4; global rect5; global rect6
    global canb; global recta
    global enteri; global enterc; global enterh
    global cnici; global cnicc; global cnich
    global emaili; global emailc; global emailh
    global namei; global namec; global nameh
    global sectioni; global sectionc; global sectionh
    global voteri; global voterc; global voteh
    global cmsi; global cmsc; global cmsh

    root=Tk()
    root.geometry("550x730")
    root.minsize(550, 710)
    root.attributes("-fullscreen", True)
    root.title("Electicider Loging System")

    imgn=Image.open("Nustlogo.png")
    imgn=imgn.resize((65,65),Image.ANTIALIAS)
    nustlogo=ImageTk.PhotoImage(imgn)

    img1=Image.open("nust.png")
    img1=img1.resize((230,210),Image.ANTIALIAS)
    pic=ImageTk.PhotoImage(img1)

    imgc=Image.open("cms1c.png")
    imgc=imgc.resize((418,50),Image.ANTIALIAS)
    cmsc=ImageTk.PhotoImage(imgc)

    img3=Image.open("cms1.png")
    img3=img3.resize((418,52),Image.ANTIALIAS)
    cmsi=ImageTk.PhotoImage(img3)

    imgc=Image.open("namec.png")
    imgc=imgc.resize((417,53),Image.ANTIALIAS)
    namec=ImageTk.PhotoImage(imgc)

    img3=Image.open("name.png")
    img3=img3.resize((418,52),Image.ANTIALIAS)
    namei=ImageTk.PhotoImage(img3)

    imgc=Image.open("voterc.png")
    imgc=imgc.resize((418,51),Image.ANTIALIAS)
    voterc=ImageTk.PhotoImage(imgc)

    img3=Image.open("voter.png")
    img3=img3.resize((418,51),Image.ANTIALIAS)
    voteri=ImageTk.PhotoImage(img3)

    imgc=Image.open("sectionc.png")
    imgc=imgc.resize((418,52),Image.ANTIALIAS)
    sectionc=ImageTk.PhotoImage(imgc)

    img3=Image.open("section.png")
    img3=img3.resize((418,52),Image.ANTIALIAS)
    sectioni=ImageTk.PhotoImage(img3)

    imgc=Image.open("emailc.png")
    imgc=imgc.resize((418,50),Image.ANTIALIAS)
    emailc=ImageTk.PhotoImage(imgc)

    img3=Image.open("email.png")
    img3=img3.resize((415,50),Image.ANTIALIAS)
    emaili=ImageTk.PhotoImage(img3)

    imgc=Image.open("cnicc.png")
    imgc=imgc.resize((418,50),Image.ANTIALIAS)
    cnicc=ImageTk.PhotoImage(imgc)

    img3=Image.open("cnic.png")
    img3=img3.resize((418,50),Image.ANTIALIAS)
    cnici=ImageTk.PhotoImage(img3)

    img5=Image.open("enter.png")
    img5=img5.resize((210,45),Image.ANTIALIAS)
    enteri=ImageTk.PhotoImage(img5)

    img6=Image.open("enterc.png")
    img6=img6.resize((210,51),Image.ANTIALIAS)
    enterc=ImageTk.PhotoImage(img6)

    img7=Image.open("enterh.png")
    img7=img7.resize((210,53),Image.ANTIALIAS)
    enterh=ImageTk.PhotoImage(img7)

    frame=Frame(root,bg="white",height=600,width=500)
    frame.place(relx=0,rely=0,relheight=1,relwidth=1)

    imagef=Frame(frame,bg='white')
    imagef.place(relx=0,rely=0.085,relwidth=1,relheight=0.29)
    textf=Frame(frame,bg="white")
    textf.place(relx=0,rely=0.38,relwidth=1,relheight=1)

    Nust=Frame(frame,bg="white")
    Nust.place(relx=0,rely=0,relwidth=0.2,relheight=0.1)

    logoN=Label(Nust,image=nustlogo,bg="white")
    logoN.place(relx=0.005,rely=0,relheight=1)

    entries=Frame(textf,bg="white")
    entries.place(relx=0,rely=0,relwidth=1,relheight=1)
    entries1=Frame(entries,bg="white")
    entries1.place(relx=0,rely=0,relwidth=1,relheight=1)

    namef=Frame(entries1,bg="white")
    namef.place(relx=0,rely=0.025,relwidth=1,relheight=0.09)
    namee=Frame(entries1,bg="white")
    namee.place(relx=0.5,rely=0.0431,width=195,relheight=0.04)

    cmsf=Frame(entries1,bg="white")
    cmsf.place(relx=0,rely=0.0968,relwidth=1,relheight=0.09)
    cmse=Frame(entries1,bg="white")
    cmse.place(relx=0.5,rely=0.1135,width=195,relheight=0.04)

    cnicf=Frame(entries1,bg="white")
    cnicf.place(relx=0,rely=0.176,relwidth=1,relheight=0.09)
    cnice=Frame(entries1,bg="white")
    cnice.place(relx=0.5,rely=0.192,width=195,relheight=0.04)

    sectf=Frame(entries1,bg="white")
    sectf.place(relx=0,rely=0.2493,relwidth=1,relheight=0.09)
    secte=Frame(entries1,bg="white")
    secte.place(relx=0.5,rely=0.266,width=195,relheight=0.04)

    voterf=Frame(entries1,bg="white")
    voterf.place(relx=0,rely=0.3236,relwidth=1,relheight=0.09)
    votere=Frame(entries1,bg="white")
    votere.place(relx=0.5,rely=0.3415,width=195,relheight=0.04)

    emailf=Frame(entries1,bg="white")
    emailf.place(relx=0,rely=0.4,relwidth=1,relheight=0.09)
    emaile=Frame(entries1,bg="white")
    emaile.place(relx=0.5,rely=0.416,width=195,relheight=0.04)

    logoN=Label(Nust,image=nustlogo,bg="white")
    logoN.place(relx=0.005,rely=0,relheight=1)

    lab=Label(imagef,bg='white',image=pic)
    lab.place(relx=0.25,rely=0,relheight=1,relwidth=0.5)
    global Name,cms,nic,sect,voter,mail
    Name=StringVar()
    cms=IntVar()
    nic=IntVar()
    sect=StringVar()
    voter=IntVar()
    mail=StringVar()

    wname=StringVar(root,value="Enter your name here")
    wcms=StringVar(root,value="Enter your Cms ID here")
    wnic=StringVar(root,value="Enter your CNIC no. here")
    wsect=StringVar(root,value="Enter your Section here")
    wvoter=StringVar(root,value="Enter your Voter ID here")
    wmail=StringVar(root,value="Enter your email here")


    lab1=Label(namef,image=namei,anchor=N,bg="white")
    entry1=Entry(namee,textvariable=wname,relief="flat",bg="white",fg="#002577",font="Garamond 11")
    entry1.pack(fill="both",expand="true")
    lab1.place(relx=0,y=0,height=400,relwidth=1)


    lab2=Label(cmsf,image=cmsi,anchor=N,bg="white")
    entry2=Entry(cmse,textvariable=wcms,relief="flat",bg="white",fg="#002577",font="Garamond 11")
    entry2.pack(fill="both",expand="true")
    lab2.place(relx=0,y=0,height=400,relwidth=1)

    lab3=Label(cnicf,image=cnici,anchor=N,bg="white")
    entry3=Entry(cnice,textvariable=wnic,relief="flat",bg="white",fg="#002577",font="Garamond 11")
    entry3.pack(fill="both",expand="true")
    lab3.place(relx=0,y=0,height=400,relwidth=1)

    lab4=Label(sectf,image=sectioni,anchor=N,bg="white")
    entry4=Entry(secte,textvariable=wsect,relief="flat",bg="white",fg="#002577",font="Garamond 11")
    entry4.pack(fill="both",expand="true")
    lab4.place(relx=0,y=0,height=400,relwidth=1)

    lab5=Label(voterf,image=voteri,anchor=N,bg="white")
    lab5.place(relx=0,y=0,height=400,relwidth=1)
    entry5=Entry(votere,textvariable=wvoter,relief="flat",bg="white",fg="#002577",font="Garamond 11")
    entry5.pack(fill="both",expand="true")

    lab6=Label(emailf,image=emaili,anchor=N,bg="white")
    lab6.place(relx=0,y=0,height=400,relwidth=1)
    entry6=Entry(emaile,textvariable=wmail,relief="flat",bg="white",fg="#002577",font="Garamond 11")
    entry6.pack(fill="both",expand="true")


    can1 = Canvas(entries1, width=10, height=40, bg="white", bd=-2, relief="flat")
    can1.place(relx=0, rely=0.036)
    rect1 = can1.create_rectangle(-20, 0, -5, 50, fill="#002577")

    can2 = Canvas(entries1, width=10, height=40, bg="white", bd=-2, relief="flat")
    can2.place(relx=0, rely=0.1125)
    rect2 = can2.create_rectangle(-20, 0, -5, 50, fill="#002577")

    can3 = Canvas(entries1, width=10, height=40, bg="white", bd=-2, relief="flat")
    can3.place(relx=0, rely=0.188)
    rect3 = can3.create_rectangle(-20, 0, -5, 50, fill="#002577")

    can4 = Canvas(entries1, width=10, height=40, bg="white", bd=-2, relief="flat")
    can4.place(relx=0, rely=0.2615)
    rect4 = can4.create_rectangle(-20, 0, -5, 50, fill="#002577")

    can5 = Canvas(entries1, width=10, height=40, bg="white", bd=-2, relief="flat")
    can5.place(relx=0, rely=0.335)
    rect5 = can5.create_rectangle(-20, 0, -5, 50, fill="#002577")

    can6 = Canvas(entries1, width=10, height=40, bg="white", bd=-2, relief="flat")
    can6.place(relx=0, rely=0.411)
    rect6 = can6.create_rectangle(-20, 0, -5, 50, fill="#002577")

    entry1.bind('<Button>',click1)
    entry1.bind('<Enter>',hoveron1)
    entry1.bind('<Leave>',hoveroff1)

    entry2.bind('<Button>',click2)
    entry2.bind('<Enter>',hoveron2)
    entry2.bind('<Leave>',hoveroff2)

    entry3.bind('<Button>',click3)
    entry3.bind('<Enter>',hoveron3)
    entry3.bind('<Leave>',hoveroff3)

    entry4.bind('<Button>',click4)
    entry4.bind('<Enter>',hoveron4)
    entry4.bind('<Leave>',hoveroff4)


    entry5.bind('<Button>',click5)
    entry5.bind('<Enter>',hoveron5)
    entry5.bind('<Leave>',hoveroff5)


    entry6.bind('<Button>',click6)
    entry6.bind('<Enter>',hoveron6)
    entry6.bind('<Leave>',hoveroff6)

    button1=Button(entries,bg='white',relief="flat",borderwidth=0,image=enteri,activebackground="white",command=checker)
    button1.place(relx=0.422,rely=0.485,width=210,relheight=0.08)


    button1.bind('<Button>',clicke)
    button1.bind('<ButtonRelease>',releasee)
    button1.bind('<Enter>',hoverone4)
    button1.bind('<Leave>',hoveroffe4)



    statusbar=Label(root,text='Please caste your vote wisely',bd=0,relief="flat",bg="#002577",fg="white",font="Helvetica 10 bold")
    statusbar.place(relx=0,rely=0.96,relheight=0.04,relwidth=1)


    canb = Canvas(frame, width=1550, height=15, bg="white", bd=-2, relief="flat")
    canb.place(x=80, y=27)
    recta = canb.create_rectangle(3000, 5, 1100, 10, fill="#002577")
    line()
    root.mainloop()
mains()